# -*- coding: utf-8 -*-
"""검진기관 자체 DB 업로드용 샘플 엑셀 양식 생성.

컬럼명은 lib/file_import.py의 FIELD_KEYWORDS / EXAM_TYPE_COLUMN_TO_API_FIELD와
정확히 일치해야 자동 인식됩니다.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT_NAME = "Arial"

HEADER_FILL = PatternFill(start_color="1F4E3D", end_color="1F4E3D", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True)
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NORMAL_FONT = Font(name=FONT_NAME)

COLUMNS = [
    ("검진기관명", True, "예: 경북대학교병원"),
    ("소재지주소", True, "예: 대구광역시 중구 동덕로 130"),
    ("전화번호", False, "예: 053-200-5114"),
    ("기관종별", False, "예: 상급종합병원 / 종합병원 / 병원 / 의원"),
    ("시도명", False, "예: 대구광역시 (비워두면 주소에서 자동으로 추정합니다)"),
    ("시군구명", False, "예: 중구 (비워두면 주소에서 자동으로 추정합니다)"),
    ("위도", False, "예: 35.8688 (소수 좌표, 지도 표시에 사용. 모르면 비워두세요)"),
    ("경도", False, "예: 128.6046"),
    ("일반건강검진", False, "Y 또는 N"),
    ("영유아검진", False, "Y 또는 N"),
    ("위암검진", False, "Y 또는 N"),
    ("간암검진", False, "Y 또는 N"),
    ("대장암검진", False, "Y 또는 N"),
    ("유방암검진", False, "Y 또는 N"),
    ("자궁경부암검진", False, "Y 또는 N"),
    ("구강검진", False, "Y 또는 N"),
    ("폐암검진", False, "Y 또는 N"),
]

EXAMPLE_ROW = [
    "경북대학교병원",
    "대구광역시 중구 동덕로 130",
    "053-200-5114",
    "종합병원",
    "대구광역시",
    "중구",
    35.8688,
    128.6046,
    "Y",
    "N",
    "Y",
    "Y",
    "Y",
    "Y",
    "N",
    "Y",
    "N",
]

wb = Workbook()

# ---------------------------------------------------------------------------
# 1. 작성양식 시트
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "작성양식"

for col_idx, (header, required, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if required:
        cell.value = f"{header} *"
    ws.column_dimensions[get_column_letter(col_idx)].width = 18

for col_idx, value in enumerate(EXAMPLE_ROW, start=1):
    cell = ws.cell(row=2, column=col_idx, value=value)
    cell.font = NORMAL_FONT
    if COLUMNS[col_idx - 1][1]:
        cell.fill = REQUIRED_FILL

ws.freeze_panes = "A2"

# Y/N 데이터 검증 (검진종류 컬럼들)
yn_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ws.add_data_validation(yn_validation)
exam_col_start = 9  # "일반건강검진"이 9번째 컬럼
exam_col_end = len(COLUMNS)
for col_idx in range(exam_col_start, exam_col_end + 1):
    col_letter = get_column_letter(col_idx)
    yn_validation.add(f"{col_letter}3:{col_letter}1000")

# ---------------------------------------------------------------------------
# 2. 작성안내 시트
# ---------------------------------------------------------------------------
guide = wb.create_sheet("작성안내")
guide.column_dimensions["A"].width = 18
guide.column_dimensions["B"].width = 14
guide.column_dimensions["C"].width = 55

guide.cell(row=1, column=1, value="컬럼명").font = HEADER_FONT
guide.cell(row=1, column=2, value="필수여부").font = HEADER_FONT
guide.cell(row=1, column=3, value="설명").font = HEADER_FONT
for c in range(1, 4):
    guide.cell(row=1, column=c).fill = HEADER_FILL
    guide.cell(row=1, column=c).alignment = Alignment(horizontal="center")

for row_idx, (header, required, desc) in enumerate(COLUMNS, start=2):
    guide.cell(row=row_idx, column=1, value=header).font = NORMAL_FONT
    guide.cell(row=row_idx, column=2, value="필수" if required else "선택").font = NORMAL_FONT
    guide.cell(row=row_idx, column=3, value=desc).font = NORMAL_FONT
    if required:
        for c in range(1, 4):
            guide.cell(row=row_idx, column=c).fill = REQUIRED_FILL

note_row = len(COLUMNS) + 3
notes = [
    "※ 노란색 배경 = 필수 입력 항목 (검진기관명, 소재지주소).",
    "※ '시도명'/'시군구명'을 채우면 주소 자동추정보다 우선 적용되어 더 정확합니다.",
    "※ 위도/경도를 모르면 비워두세요 — 그래도 등록은 되고, 지도에만 안 나타납니다.",
    "※ 검진종류 컬럼은 Y 또는 N만 입력하세요 (드롭다운으로 선택 가능하게 해뒀습니다).",
    "※ 이 파일을 다 채운 뒤 '전체 데이터 파일 업로드' 페이지에 그대로 업로드하시면 됩니다.",
    "※ 1행은 헤더, 2행은 예시이니 실제 데이터는 3행부터 입력해주세요 (2행 예시는 지우셔도 됩니다).",
]
for i, note in enumerate(notes):
    cell = guide.cell(row=note_row + i, column=1, value=note)
    cell.font = Font(name=FONT_NAME, italic=True, color="555555")
    guide.merge_cells(start_row=note_row + i, start_column=1, end_row=note_row + i, end_column=3)

wb.save("/mnt/user-data/outputs/검진기관_자체DB_업로드양식.xlsx")
print("저장 완료")
