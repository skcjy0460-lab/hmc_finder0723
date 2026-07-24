# -*- coding: utf-8 -*-
"""
검진기관 데이터 로컬 SQLite 캐시.

getRegnHmcList(지역별 조회)가 지역 파라미터로는 정상 동작하지 않는 것으로 확인되어,
getHchkTypesHmcList(전국조회)로 전체 데이터를 한 번에 받아와 로컬 DB에 저장하고
화면의 검색/필터는 이 DB에서 처리합니다.
"""
from __future__ import annotations

import datetime as dt
import sqlite3
from pathlib import Path

import pandas as pd

from . import coord_utils

DB_PATH = Path(__file__).resolve().parent.parent / "data" / "hmc_cache.db"

# 검진 항목별 "가능" 여부 코드 컬럼 매핑 (한글 라벨 -> API 응답 필드명)
# ⚠️ 실제 코드값(예: "1"/"Y"/"가능")이 무엇인지 아직 확인 전이라
#    CHRG_AVAILABLE_VALUES에 자주 쓰이는 값들을 넉넉히 넣어두었습니다.
#    첫 실제 응답을 받으면 print(df[col].unique())로 확인 후 필요시 값만 추가하세요.
EXAM_TYPE_FIELDS = {
    "일반건강검진": "gren_chrg_type_cd",
    "영유아검진": "ichk_chrg_type_cd",
    "생애전환기검진": "mchk_chrg_type_cd",
    "위암검진": "stmca_exmd_chrg_type_cd",
    "간암검진": "lvca_exmd_chrg_type_cd",
    "대장암검진": "cc_exmd_chrg_type_cd",
    "유방암검진": "bc_exmd_chrg_type_cd",
    "자궁경부암검진": "cvxca_exmd_chrg_type_cd",
}
CHRG_AVAILABLE_VALUES = {"1", "Y", "y", "가능", "01", "true", "True"}

SCHEMA = """
CREATE TABLE IF NOT EXISTS hmc (
    hmc_no TEXT PRIMARY KEY,
    hmc_nm TEXT,
    si_do_cd TEXT,
    si_gun_gu_cd TEXT,
    loc_addr TEXT,
    loc_post_no TEXT,
    hmc_tel_no TEXT,
    exmdr_tel_no TEXT,
    exmdr_fax_no TEXT,
    ykindnm TEXT,
    cx_vl TEXT,
    cy_vl TEXT,
    lat REAL,
    lng REAL,
    gren_chrg_type_cd TEXT,
    ichk_chrg_type_cd TEXT,
    mchk_chrg_type_cd TEXT,
    stmca_exmd_chrg_type_cd TEXT,
    lvca_exmd_chrg_type_cd TEXT,
    cc_exmd_chrg_type_cd TEXT,
    bc_exmd_chrg_type_cd TEXT,
    cvxca_exmd_chrg_type_cd TEXT,
    sync_region_full_cd TEXT,
    synced_at TEXT
);

CREATE TABLE IF NOT EXISTS sync_log (
    si_gun_gu_cd TEXT PRIMARY KEY,
    si_do_cd TEXT,
    synced_at TEXT,
    row_count INTEGER
);
"""


def _connect() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)
    return conn


def _row_from_api_item(item: dict) -> dict:
    """API 응답 item(카멜케이스)을 DB 컬럼(스네이크케이스)으로 매핑."""
    cx_vl = item.get("cxVl") or item.get("cxV1")  # 캡처본 OCR 오차(V1) 대비 폴백
    cy_vl = item.get("cyVl") or item.get("cyV1")
    latlng = coord_utils.to_wgs84(cx_vl, cy_vl)
    lat, lng = latlng if latlng else (None, None)

    return {
        "hmc_no": item.get("hmcNo"),
        "hmc_nm": item.get("hmcNm"),
        "si_do_cd": item.get("siDoCd"),
        "si_gun_gu_cd": item.get("siGunGuCd"),
        "loc_addr": item.get("locAddr"),
        "loc_post_no": item.get("locPostNo"),
        "hmc_tel_no": item.get("hmcTelNo"),
        "exmdr_tel_no": item.get("exmdrTelNo"),
        "exmdr_fax_no": item.get("exmdrFaxNo"),
        "ykindnm": item.get("ykindnm"),
        "cx_vl": cx_vl,
        "cy_vl": cy_vl,
        "lat": lat,
        "lng": lng,
        "gren_chrg_type_cd": item.get("grenChrgTypeCd"),
        "ichk_chrg_type_cd": item.get("ichkChrgTypeCd"),
        "mchk_chrg_type_cd": item.get("mchkChrgTypeCd"),
        "stmca_exmd_chrg_type_cd": item.get("stmcaExmdChrgTypeCd"),
        # API 원본에 lvcaExmdChrgType(코드 접미사 Cd 누락) 오탈자가 있어 둘 다 대응
        "lvca_exmd_chrg_type_cd": item.get("lvcaExmdChrgTypeCd") or item.get("lvcaExmdChrgType"),
        "cc_exmd_chrg_type_cd": item.get("ccExmdChrgTypeCd"),
        "bc_exmd_chrg_type_cd": item.get("bcExmdChrgTypeCd"),
        "cvxca_exmd_chrg_type_cd": item.get("cvxcaExmdChrgTypeCd"),
        "synced_at": dt.datetime.now().isoformat(timespec="seconds"),
    }


def upsert_items(items: list[dict], sync_key: str = "ALL_NATIONWIDE") -> int:
    """API에서 받은 검진기관 item 리스트를 DB에 upsert하고 sync_log를 갱신.

    getRegnHmcList(지역별 조회)가 지역 파라미터로는 데이터를 반환하지 않는 것으로
    확인되어(2026-07-23), 전국조회 API로 한 번에 받아온 결과를 저장하는 방식으로
    전환했습니다. 지역 필터링은 각 item에 이미 포함된 siDoCd/siGunGuCd 값을
    그대로 저장해두고 search_local()에서 처리합니다.
    """
    if not items:
        return 0

    rows = [_row_from_api_item(it) for it in items]
    cols = list(rows[0].keys())
    placeholders = ", ".join(f":{c}" for c in cols)
    col_list = ", ".join(cols)
    update_clause = ", ".join(f"{c}=excluded.{c}" for c in cols if c != "hmc_no")

    sql = f"""
    INSERT INTO hmc ({col_list}) VALUES ({placeholders})
    ON CONFLICT(hmc_no) DO UPDATE SET {update_clause}
    """

    conn = _connect()
    try:
        with conn:
            conn.executemany(sql, rows)
            conn.execute(
                """
                INSERT INTO sync_log (si_gun_gu_cd, si_do_cd, synced_at, row_count)
                VALUES (?, ?, ?, ?)
                ON CONFLICT(si_gun_gu_cd) DO UPDATE SET
                    synced_at=excluded.synced_at,
                    row_count=excluded.row_count
                """,
                (sync_key, None, dt.datetime.now().isoformat(timespec="seconds"), len(rows)),
            )
    finally:
        conn.close()
    return len(rows)


def get_last_sync(sync_key: str = "ALL_NATIONWIDE") -> dict | None:
    conn = _connect()
    try:
        cur = conn.execute(
            "SELECT si_gun_gu_cd, si_do_cd, synced_at, row_count FROM sync_log WHERE si_gun_gu_cd=?",
            (sync_key,),
        )
        row = cur.fetchone()
        if not row:
            return None
        keys = ["si_gun_gu_cd", "si_do_cd", "synced_at", "row_count"]
        return dict(zip(keys, row))
    finally:
        conn.close()


def search_local(
    si_do_cd: str | int | None = None,
    si_gun_gu_cd: str | int | None = None,
    hmc_nm_keyword: str | None = None,
    required_exam_types: list[str] | None = None,
) -> pd.DataFrame:
    """로컬 DB에서 조건에 맞는 검진기관을 조회합니다.

    si_do_cd / si_gun_gu_cd: CodeServices(getSiDoList/getSiGunGuList)가 반환하는
    분리형 코드값을 그대로 넘겨주세요. 전국조회 API의 item 자체에 들어있는
    siDoCd/siGunGuCd와 같은 코드 체계라 그대로 매칭됩니다.
    """
    conn = _connect()
    try:
        df = pd.read_sql_query("SELECT * FROM hmc", conn)
    finally:
        conn.close()

    if df.empty:
        return df

    if si_do_cd not in (None, ""):
        df = df[df["si_do_cd"].astype(str) == str(si_do_cd)]
    if si_gun_gu_cd not in (None, ""):
        df = df[df["si_gun_gu_cd"].astype(str) == str(si_gun_gu_cd)]
    if hmc_nm_keyword:
        df = df[df["hmc_nm"].fillna("").str.contains(hmc_nm_keyword, case=False, na=False)]

    if required_exam_types:
        for label in required_exam_types:
            field = EXAM_TYPE_FIELDS.get(label)
            if field and field in df.columns:
                df = df[df[field].isin(CHRG_AVAILABLE_VALUES)]

    return df.reset_index(drop=True)


def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """검색 결과 DataFrame을 보기 좋은 스타일의 엑셀 바이트로 변환."""
    import io

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    display_cols = {
        "hmc_nm": "기관명",
        "ykindnm": "기관종별",
        "loc_addr": "주소",
        "hmc_tel_no": "전화번호",
        "exmdr_tel_no": "검진문의전화",
    }
    export_df = df[[c for c in display_cols if c in df.columns]].rename(columns=display_cols)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="검진기관")
        ws = writer.sheets["검진기관"]

        header_fill = PatternFill(start_color="1F4E3D", end_color="1F4E3D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, _ in enumerate(export_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    return buffer.getvalue()
